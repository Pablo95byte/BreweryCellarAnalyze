#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# App: Analisi Tank – per Tank + per Material + Totale (no SG) + Selezione Giorno Singolo
# Autore: M365 Copilot
# Requisiti: Python 3.8+ (tkinter). Per export XLSX facoltativo: openpyxl.
# Regole:
#  - Average Gravity == Average Plato (usati come "Gravity" così come sono dal CSV).
#  - f(A) = ((0.0000188792 * G + 0.003646886) * G + 1.001077) * G - 0.01223565
#  - Kg estratto (riga) = f(A) * Level; aggregare sul periodo filtrato.
#  - Material mapping: 7=ichnusa, 8=non filtrata, 9=cruda, 28=ambra limpida. Altri valori mantenuti com'è (incluso '0').

import csv
import os
import re
import math
from datetime import datetime, timedelta
from collections import defaultdict

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import Workbook
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    _HAS_MATPLOTLIB = True
except Exception:
    _HAS_MATPLOTLIB = False

APP_TITLE = "Analisi Tank – per Tank + per Material + Totale (no SG)"
APP_VERSION = "1.0"
APP_AUTHOR = "PA"
APP_EMAIL = "PA@HE.IT"
APP_DEPT = "ASS_ST"

# ---------------------- Utility ----------------------

def to_float(s):
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def fmt_it(x, nd=2):
    if x is None:
        return ""
    s = f"{x:,.{nd}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def parse_time(s):
    if s is None:
        return None
    s = s.strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None

# f(A) con A = Gravity (dato dal CSV)
# f(A) = ((0.0000188792 * G + 0.003646886) * G + 1.001077) * G - 0.01223565

def fA(gravity):
    if gravity is None:
        return None
    return ((0.0000188792*gravity + 0.003646886)*gravity + 1.001077)*gravity - 0.01223565

# Material mapping helpers
_MAP_CODES = {
    '7': 'ichnusa',
    '8': 'non filtrata',
    '9': 'cruda',
    '28': 'ambra limpida',
    '0': 'vuoto',
    '10': 'ich(prop)',
    
}

_DEF_EMPTY = '(vuoto)'


def normalize_material(val):
    if val is None:
        return _DEF_EMPTY
    s = str(val).strip()
    if not s:
        return _DEF_EMPTY
    # prova come codice
    if s in _MAP_CODES:
        return _MAP_CODES[s]
    # prova numerico (include 0)
    try:
        k = str(int(float(s)))
        if k in _MAP_CODES:
            return _MAP_CODES[k]
        return k
    except Exception:
        pass
    # prova testo
    low = s.lower().strip()
    # normalizza accentazione semplice
    repl = {
        'ì': 'i', 'í': 'i', 'ï': 'i', 'î': 'i',
        'à': 'a', 'á': 'a', 'ä': 'a', 'â': 'a',
        'è': 'e', 'é': 'e', 'ë': 'e', 'ê': 'e',
        'ò': 'o', 'ó': 'o', 'ö': 'o', 'ô': 'o',
        'ù': 'u', 'ú': 'u', 'ü': 'u', 'û': 'u',
    }
    low = ''.join(repl.get(ch, ch) for ch in low)
    if 'ichnusa' in low:
        return 'ichnusa'
    if 'non filtrata' in low or 'nonfiltrata' in low:
        return 'non filtrata'
    if 'cruda' in low:
        return 'cruda'
    if 'ambra' in low and 'limpida' in low:
        return 'ambra limpida'
    return s


# ---------------------- Analyzer ----------------------
class Analyzer:
    # Accetta sia "Average Plato" sia "Average Gravity" come stessi dati
    RE_AVG   = re.compile(r"^(FST|BBT)\s*([0-9]+)\s*Average\s*(Plato|Gravity)$", re.I)
    RE_LEVEL = re.compile(r"^(FST|BBT)\s*([0-9]+)\s*Level\s*$", re.I)
    RE_MAT   = re.compile(r"^(FST|BBT)\s*([0-9]+)\s*Material\s*$", re.I)

    def __init__(self, path):
        self.path = path
        self.header = []
        self.rows = []
        self.time_idx = None
        self.avg_cols = []    # (idx, tank_key, family)
        self.level_idx = {}   # tank_key -> idx
        self.material_idx = {}# tank_key -> idx
        self.min_time = None
        self.max_time = None
        self._load()

    def _load(self):
        with open(self.path, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            raise ValueError("CSV vuoto")
        self.header = [h.strip() for h in rows[0]]
        self.rows = rows[1:]

        # Time
        for i, h in enumerate(self.header):
            if h.strip().lower() == 'time':
                self.time_idx = i
                break

        # Average / Level / Material
        for idx, col in enumerate(self.header):
            m = self.RE_AVG.match(col)
            if m:
                fam = m.group(1).upper()
                num = m.group(2)
                tank_key = f"{fam}{num}"
                self.avg_cols.append((idx, tank_key, fam))
            m2 = self.RE_LEVEL.match(col)
            if m2:
                fam2 = m2.group(1).upper()
                num2 = m2.group(2)
                self.level_idx[f"{fam2}{num2}"] = idx
            m3 = self.RE_MAT.match(col)
            if m3:
                fam3 = m3.group(1).upper()
                num3 = m3.group(2)
                self.material_idx[f"{fam3}{num3}"] = idx

        # range date
        if self.time_idx is not None:
            for r in self.rows:
                if self.time_idx < len(r):
                    dt = parse_time(r[self.time_idx])
                    if dt:
                        self.min_time = dt if self.min_time is None or dt < self.min_time else self.min_time
                        self.max_time = dt if self.max_time is None or dt > self.max_time else self.max_time

    def analyze(self, t_from=None, t_to=None, include_fst=True, include_bbt=True):
        # Per tank (aggregato): Gravity_last, Volume_last, Material_last, sum_fA, sum_kg, count
        by_tank = {}
        # Per materiale (aggregato su tutte le righe nel periodo): sum_kg, sum_fA, count
        by_material = {}
        # Debug: lista di tuple (timestamp, tank, material, gravity, level, fA_val, kg_ext)
        debug_data = []

        def tank_rec(tk):
            r = by_tank.get(tk)
            if not r:
                r = {'G_last': None, 'V_last': None, 'M_last': None, 't_last': None, 'sum_fA': 0.0, 'sum_kg': 0.0, 'count': 0}
                by_tank[tk] = r
            return r

        def mat_rec(name):
            name = normalize_material(name)
            r = by_material.get(name)
            if not r:
                r = {'sum_kg': 0.0, 'sum_fA': 0.0, 'count': 0}
                by_material[name] = r
            return r

        for row in self.rows:
            # filtro temporale
            if self.time_idx is not None and (t_from or t_to):
                dt = parse_time(row[self.time_idx] if self.time_idx < len(row) else None)
                if dt is None:
                    continue
                if t_from and dt < t_from:
                    continue
                if t_to and dt > t_to:
                    continue
            else:
                dt = None

            for idx, tank_key, fam in self.avg_cols:
                if fam == 'FST' and not include_fst:
                    continue
                if fam == 'BBT' and not include_bbt:
                    continue
                if idx >= len(row):
                    continue
                G = to_float(row[idx])
                if G is None or (isinstance(G, float) and math.isnan(G)):
                    continue
                fval = fA(G)
                if fval is None or (isinstance(fval, float) and math.isnan(fval)):
                    continue

                # Level / Volume corrente
                Lidx = self.level_idx.get(tank_key)
                V = to_float(row[Lidx]) if (Lidx is not None and Lidx < len(row)) else None
                if V is None or (isinstance(V, float) and math.isnan(V)):
                    V = 0.0
                if V < 0:
                    V = 0.0

                # Material corrente
                midx = self.material_idx.get(tank_key)
                M = row[midx] if (midx is not None and midx < len(row)) else None
                M_norm = normalize_material(M)

                kg_ext_i = fval * V

                # Aggiungi ai dati di debug
                debug_data.append((dt, tank_key, M_norm, G, V, fval, kg_ext_i))

                # aggregati tank
                tr = tank_rec(tank_key)
                if dt is not None:
                    if tr['t_last'] is None or dt > tr['t_last']:
                        tr['t_last'] = dt
                        tr['G_last'] = G
                        tr['V_last'] = V
                        tr['M_last'] = M_norm
                else:
                    tr['G_last'] = G
                    tr['V_last'] = V
                    tr['M_last'] = M_norm
                tr['sum_fA'] += fval
                tr['sum_kg'] += kg_ext_i
                tr['count'] += 1

                # aggregati materiale
                mr = mat_rec(M_norm)
                mr['sum_kg'] += kg_ext_i
                mr['sum_fA'] += fval
                mr['count']  += 1

        # output ordinati
        tank_rows = []
        for t, s in by_tank.items():
            tank_rows.append((t, s['M_last'], s['G_last'], s['V_last'], s['sum_fA'], s['sum_kg'], s['count']))
        tank_rows.sort(key=lambda x: (-x[5], x[0]))  # per kg desc

        mat_rows = []
        for m, s in by_material.items():
            mat_rows.append((m, s['sum_kg'], s['sum_fA'], s['count']))
        mat_rows.sort(key=lambda x: (-x[1], x[0]))  # per kg desc

        # Ordina debug_data per timestamp
        debug_data.sort(key=lambda x: x[0] if x[0] else datetime.min)

        return tank_rows, mat_rows, debug_data

    def analyze_all_days(self, include_fst=True, include_bbt=True):
        """Analizza tutti i giorni disponibili per i grafici temporali"""
        if self.time_idx is None:
            return {}
        
        # Raggruppa per giorno
        daily_data = defaultdict(lambda: {'kg': 0.0, 'by_material': defaultdict(float), 'by_tank': defaultdict(float)})
        
        for row in self.rows:
            dt = parse_time(row[self.time_idx] if self.time_idx < len(row) else None)
            if dt is None:
                continue
            
            day_key = dt.date().strftime("%Y-%m-%d")
            
            for idx, tank_key, fam in self.avg_cols:
                if fam == 'FST' and not include_fst:
                    continue
                if fam == 'BBT' and not include_bbt:
                    continue
                if idx >= len(row):
                    continue
                G = to_float(row[idx])
                if G is None or (isinstance(G, float) and math.isnan(G)):
                    continue
                fval = fA(G)
                if fval is None or (isinstance(fval, float) and math.isnan(fval)):
                    continue

                Lidx = self.level_idx.get(tank_key)
                V = to_float(row[Lidx]) if (Lidx is not None and Lidx < len(row)) else None
                if V is None or (isinstance(V, float) and math.isnan(V)):
                    V = 0.0
                if V < 0:
                    V = 0.0

                midx = self.material_idx.get(tank_key)
                M = row[midx] if (midx is not None and midx < len(row)) else None
                M_norm = normalize_material(M)

                kg_ext_i = fval * V
                
                daily_data[day_key]['kg'] += kg_ext_i
                daily_data[day_key]['by_material'][M_norm] += kg_ext_i
                daily_data[day_key]['by_tank'][tank_key] += kg_ext_i
        
        return daily_data


# ---------------------- GUI ----------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x760")
        self.minsize(1180, 700)
        self._cache_tank = []
        self._cache_mat = []
        self._cache_debug = []
        self.an = None
        self.current_file = None
        self.var_exclude_mat0 = tk.BooleanVar(value=False)
        self._tot_win = None
        # Single-day state
        self.sel_day = tk.StringVar(value="")
        self.days_list = []
        
        # Mostra splash screen
        self.show_splash()
        
        self._build()

    def show_splash(self):
        """Mostra splash screen all'avvio"""
        splash = tk.Toplevel(self)
        splash.title("")
        splash.overrideredirect(True)
        
        # Centra la finestra
        w, h = 400, 250
        x = (splash.winfo_screenwidth() // 2) - (w // 2)
        y = (splash.winfo_screenheight() // 2) - (h // 2)
        splash.geometry(f"{w}x{h}+{x}+{y}")
        
        # Contenuto
        frame = ttk.Frame(splash, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=APP_TITLE, font=("Segoe UI", 16, "bold")).pack(pady=(20,10))
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        ttk.Label(frame, text=f"Versione {APP_VERSION}", font=("Segoe UI", 11)).pack(pady=5)
        ttk.Label(frame, text=f"Sviluppato da {APP_AUTHOR}", font=("Segoe UI", 10)).pack(pady=5)
        ttk.Label(frame, text=APP_DEPT, font=("Segoe UI", 9), foreground="#666").pack(pady=2)
        ttk.Label(frame, text=APP_EMAIL, font=("Segoe UI", 9), foreground="#0066cc").pack(pady=2)
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        ttk.Label(frame, text="Caricamento in corso...", font=("Segoe UI", 9), foreground="#999").pack(pady=(10,20))
        
        splash.update()
        
        # Chiudi dopo 2.5 secondi
        self.after(2500, splash.destroy)

    def _build(self):
        # Menu bar
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
        
        top = ttk.Frame(self)
        top.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(top, text="Apri CSV...", command=self.on_open).pack(side=tk.LEFT)
        self.lbl_file = ttk.Label(top, text="Nessun file caricato", foreground="#555")
        self.lbl_file.pack(side=tk.LEFT, padx=10)

        filt = ttk.Frame(self)
        filt.pack(fill=tk.X, padx=10, pady=(0,8))
        self.b_fst = tk.BooleanVar(value=True)
        self.b_bbt = tk.BooleanVar(value=True)
        ttk.Checkbutton(filt, text="FST", variable=self.b_fst).pack(side=tk.LEFT, padx=(0,10))
        ttk.Checkbutton(filt, text="BBT", variable=self.b_bbt).pack(side=tk.LEFT, padx=(0,10))
        ttk.Button(filt, text="Applica", command=self.on_apply).pack(side=tk.LEFT)

        # --- Selettore Giorno Singolo ---
        single_frame = ttk.Frame(self)
        single_frame.pack(fill=tk.X, padx=10, pady=(0,8))
        ttk.Button(single_frame, text="◀", width=2, command=self.on_prev_day).pack(side=tk.LEFT, padx=(0,5))
        ttk.Label(single_frame, text="Giorno:").pack(side=tk.LEFT)
        self.cb_day = ttk.Combobox(single_frame, textvariable=self.sel_day, width=12, state="readonly")
        self.cb_day.pack(side=tk.LEFT, padx=(5,5))
        self.cb_day.bind("<<ComboboxSelected>>", lambda e: self.on_select_day())
        ttk.Button(single_frame, text="▶", width=2, command=self.on_next_day).pack(side=tk.LEFT)

        # --- Notebook con tabs ---
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

        # Tab 1: Riepilogo
        self._build_summary_tab()
        
        # Tab 2: Debug Dati
        self._build_debug_tab()
        
        # Tab 3: Grafici (solo se matplotlib disponibile)
        if _HAS_MATPLOTLIB:
            self._build_charts_tab()
        
        # Tab 4: Analisi Variazioni
        self._build_variations_tab()

        actions = ttk.Frame(self)
        actions.pack(fill=tk.X, padx=10, pady=(0,10))
        ttk.Button(actions, text="Esporta per Tank (CSV)", command=self.on_export_tank_csv).pack(side=tk.LEFT)
        ttk.Button(actions, text="Esporta per Materiale (CSV)", command=self.on_export_mat_csv).pack(side=tk.LEFT, padx=(10,0))
        ttk.Button(actions, text="Esporta Debug (CSV)", command=self.on_export_debug_csv).pack(side=tk.LEFT, padx=(10,0))
        ttk.Button(actions, text="Esporta Variazioni (CSV)", command=self.on_export_variations_csv).pack(side=tk.LEFT, padx=(10,0))
        btnx = ttk.Button(actions, text="Esporta report (XLSX)", command=self.on_export_xlsx)
        if not _HAS_OPENPYXL:
            btnx.state(["disabled"])
        btnx.pack(side=tk.LEFT, padx=(10,0))

        hint = ttk.Label(self, text=(
            "Average Gravity == Average Plato (valore usato come Gravity). "
            "f(A) = ((0.0000188792*G + 0.003646886)*G + 1.001077)*G - 0.01223565; "
            "Kg estratto (riga) = f(A) * Level. Mapping: 7=ichnusa, 8=non filtrata, 9=cruda, 28=ambra limpida."
        ), foreground="#555")
        hint.pack(fill=tk.X, padx=10, pady=(0,5))
        
        # Footer
        footer = ttk.Frame(self)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer_text = f"Sviluppato da {APP_AUTHOR} ({APP_DEPT}) - v{APP_VERSION} - {APP_EMAIL}"
        ttk.Label(footer, text=footer_text, foreground="#888", font=("Segoe UI", 8)).pack(pady=5)

    def _build_summary_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Riepilogo")

        # --- Riepilogo per Materiale ---
        mat_frame = ttk.LabelFrame(tab, text="Riepilogo per Materiale")
        mat_frame.pack(fill=tk.X, padx=10, pady=10)
        cols_m = ("m", "kg", "fa", "n")
        self.tv_mat = ttk.Treeview(mat_frame, columns=cols_m, show="headings", height=6)
        self.tv_mat.heading("m", text="Materiale")
        self.tv_mat.heading("kg", text="Kg estratto (somma)")
        self.tv_mat.heading("fa", text="Somma f(A)")
        self.tv_mat.heading("n", text="N. Tank")
        self.tv_mat.column("m", width=240, anchor=tk.W)
        self.tv_mat.column("kg", width=160, anchor=tk.E)
        self.tv_mat.column("fa", width=160, anchor=tk.E)
        self.tv_mat.column("n", width=100, anchor=tk.E)
        self.tv_mat.pack(fill=tk.X, padx=6, pady=6)

        # --- Barra Totale Cantina ---
        tot_bar = ttk.Frame(tab)
        tot_bar.pack(fill=tk.X, padx=10, pady=(0,8))
        ttk.Checkbutton(tot_bar, text="Escludi Material = 0 dal totale", variable=self.var_exclude_mat0, command=self.refresh_total_window).pack(side=tk.LEFT)
        ttk.Button(tot_bar, text="Totale Cantina...", command=self.on_show_total).pack(side=tk.LEFT, padx=(10,0))

        # --- Tabella per Tank ---
        tank_frame = ttk.LabelFrame(tab, text="Per Tank")
        tank_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        cols = ("tank", "mat", "g_last", "v_last", "sum_fa", "kg_ext", "n")
        self.tv = ttk.Treeview(tank_frame, columns=cols, show="headings", height=16)
        self.tv.heading("tank", text="Tank")
        self.tv.heading("mat", text="Materiale")
        self.tv.heading("g_last", text="Gravity (plato)")
        self.tv.heading("v_last", text="Volume (hl)")
        self.tv.heading("sum_fa", text="Somma f(A)")
        self.tv.heading("kg_ext", text="Kg estratto")
        self.tv.heading("n", text="Misure")
        self.tv.column("tank", width=100, anchor=tk.W)
        self.tv.column("mat", width=200, anchor=tk.W)
        self.tv.column("g_last", width=140, anchor=tk.E)
        self.tv.column("v_last", width=140, anchor=tk.E)
        self.tv.column("sum_fa", width=160, anchor=tk.E)
        self.tv.column("kg_ext", width=140, anchor=tk.E)
        self.tv.column("n", width=80, anchor=tk.E)

        vsb = ttk.Scrollbar(tank_frame, orient="vertical", command=self.tv.yview)
        self.tv.configure(yscrollcommand=vsb.set)
        self.tv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.LEFT, fill=tk.Y)

    def _build_debug_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Debug Dati")

        info = ttk.Label(tab, text="Dettaglio di ogni riga elaborata nel giorno selezionato", foreground="#555")
        info.pack(fill=tk.X, padx=10, pady=10)

        # Tabella debug
        debug_frame = ttk.Frame(tab)
        debug_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        cols_d = ("time", "tank", "material", "gravity", "level", "fA", "kg")
        self.tv_debug = ttk.Treeview(debug_frame, columns=cols_d, show="headings", height=20)
        self.tv_debug.heading("time", text="Timestamp")
        self.tv_debug.heading("tank", text="Tank")
        self.tv_debug.heading("material", text="Materiale")
        self.tv_debug.heading("gravity", text="Gravity")
        self.tv_debug.heading("level", text="Level (hl)")
        self.tv_debug.heading("fA", text="f(A)")
        self.tv_debug.heading("kg", text="Kg estratto")
        
        self.tv_debug.column("time", width=150, anchor=tk.W)
        self.tv_debug.column("tank", width=80, anchor=tk.W)
        self.tv_debug.column("material", width=150, anchor=tk.W)
        self.tv_debug.column("gravity", width=100, anchor=tk.E)
        self.tv_debug.column("level", width=100, anchor=tk.E)
        self.tv_debug.column("fA", width=120, anchor=tk.E)
        self.tv_debug.column("kg", width=120, anchor=tk.E)

        vsb_d = ttk.Scrollbar(debug_frame, orient="vertical", command=self.tv_debug.yview)
        self.tv_debug.configure(yscrollcommand=vsb_d.set)
        self.tv_debug.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb_d.pack(side=tk.LEFT, fill=tk.Y)

        # Totale debug
        self.lbl_debug_total = ttk.Label(tab, text="Totale Kg estratto (debug): -", font=("Segoe UI", 10, "bold"))
        self.lbl_debug_total.pack(fill=tk.X, padx=10, pady=(5,10))

    def _build_charts_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Grafici")

        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(btn_frame, text="Genera Grafici", command=self.on_generate_charts).pack(side=tk.LEFT)
        ttk.Label(btn_frame, text="(analizza tutti i giorni disponibili)", foreground="#555").pack(side=tk.LEFT, padx=10)

        # Frame per i grafici
        self.charts_frame = ttk.Frame(tab)
        self.charts_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

    def _build_variations_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Analisi Variazioni")

        info = ttk.Label(tab, text="Variazioni di livello e perdite tra misurazioni successive nel giorno selezionato", foreground="#555")
        info.pack(fill=tk.X, padx=10, pady=10)

        # Frame per selezione tank
        select_frame = ttk.Frame(tab)
        select_frame.pack(fill=tk.X, padx=10, pady=(0,10))
        ttk.Label(select_frame, text="Filtra per Tank:").pack(side=tk.LEFT)
        self.var_filter_tank = tk.StringVar(value="Tutti")
        self.cb_filter_tank = ttk.Combobox(select_frame, textvariable=self.var_filter_tank, width=15, state="readonly")
        self.cb_filter_tank['values'] = ["Tutti"]
        self.cb_filter_tank.pack(side=tk.LEFT, padx=(5,10))
        self.cb_filter_tank.bind("<<ComboboxSelected>>", lambda e: self.update_variations_table())
        ttk.Button(select_frame, text="Aggiorna", command=self.update_variations_table).pack(side=tk.LEFT)

        # Tabella variazioni
        var_frame = ttk.Frame(tab)
        var_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        cols_v = ("time", "tank", "material", "level_prev", "level_curr", "delta_level", "gravity_prev", "gravity_curr", "kg_prev", "kg_curr", "delta_kg")
        self.tv_variations = ttk.Treeview(var_frame, columns=cols_v, show="headings", height=20)
        self.tv_variations.heading("time", text="Timestamp")
        self.tv_variations.heading("tank", text="Tank")
        self.tv_variations.heading("material", text="Materiale")
        self.tv_variations.heading("level_prev", text="Level Prec (hl)")
        self.tv_variations.heading("level_curr", text="Level Corr (hl)")
        self.tv_variations.heading("delta_level", text="ΔLevel (hl)")
        self.tv_variations.heading("gravity_prev", text="Gravity Prec")
        self.tv_variations.heading("gravity_curr", text="Gravity Corr")
        self.tv_variations.heading("kg_prev", text="Kg Prec")
        self.tv_variations.heading("kg_curr", text="Kg Corr")
        self.tv_variations.heading("delta_kg", text="ΔKg")
        
        self.tv_variations.column("time", width=140, anchor=tk.W)
        self.tv_variations.column("tank", width=70, anchor=tk.W)
        self.tv_variations.column("material", width=120, anchor=tk.W)
        self.tv_variations.column("level_prev", width=100, anchor=tk.E)
        self.tv_variations.column("level_curr", width=100, anchor=tk.E)
        self.tv_variations.column("delta_level", width=90, anchor=tk.E)
        self.tv_variations.column("gravity_prev", width=90, anchor=tk.E)
        self.tv_variations.column("gravity_curr", width=90, anchor=tk.E)
        self.tv_variations.column("kg_prev", width=90, anchor=tk.E)
        self.tv_variations.column("kg_curr", width=90, anchor=tk.E)
        self.tv_variations.column("delta_kg", width=90, anchor=tk.E)

        vsb_v = ttk.Scrollbar(var_frame, orient="vertical", command=self.tv_variations.yview)
        hsb_v = ttk.Scrollbar(var_frame, orient="horizontal", command=self.tv_variations.xview)
        self.tv_variations.configure(yscrollcommand=vsb_v.set, xscrollcommand=hsb_v.set)
        self.tv_variations.grid(row=0, column=0, sticky='nsew')
        vsb_v.grid(row=0, column=1, sticky='ns')
        hsb_v.grid(row=1, column=0, sticky='ew')
        var_frame.grid_rowconfigure(0, weight=1)
        var_frame.grid_columnconfigure(0, weight=1)

        # Summary variazioni
        summary_frame = ttk.LabelFrame(tab, text="Riepilogo Variazioni")
        summary_frame.pack(fill=tk.X, padx=10, pady=(0,10))
        self.lbl_var_summary = ttk.Label(summary_frame, text="Seleziona un giorno per vedere le variazioni", foreground="#555")
        self.lbl_var_summary.pack(padx=10, pady=10, anchor=tk.W)

    def show_about(self):
        """Mostra finestra About"""
        about = tk.Toplevel(self)
        about.title("About")
        about.resizable(False, False)
        about.geometry("400x300")
        
        # Centra
        about.update_idletasks()
        x = (about.winfo_screenwidth() // 2) - (about.winfo_width() // 2)
        y = (about.winfo_screenheight() // 2) - (about.winfo_height() // 2)
        about.geometry(f"+{x}+{y}")
        
        frame = ttk.Frame(about, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=APP_TITLE, font=("Segoe UI", 14, "bold")).pack(pady=(10,5))
        ttk.Label(frame, text=f"Versione {APP_VERSION}", font=("Segoe UI", 10)).pack(pady=5)
        
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        ttk.Label(frame, text="Sviluppato da:", font=("Segoe UI", 9, "bold")).pack(pady=(5,2))
        ttk.Label(frame, text=APP_AUTHOR, font=("Segoe UI", 10)).pack(pady=2)
        ttk.Label(frame, text=APP_DEPT, font=("Segoe UI", 9), foreground="#666").pack(pady=2)
        ttk.Label(frame, text=APP_EMAIL, font=("Segoe UI", 9), foreground="#0066cc").pack(pady=2)
        
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        info_text = ("Tool per l'analisi di tank FST/BBT con calcolo\n"
                    "estratto, variazioni di livello e grafici temporali.")
        ttk.Label(frame, text=info_text, font=("Segoe UI", 8), foreground="#555", justify=tk.CENTER).pack(pady=5)
        
        ttk.Button(frame, text="Chiudi", command=about.destroy).pack(pady=(15,10))

    # ---------------------- Giorno singolo ----------------------
    def populate_days(self):
        """Crea l'elenco dei giorni (YYYY-MM-DD) presenti nel CSV e popola la combo."""
        self.days_list = []
        if not self.an or not self.an.rows or self.an.time_idx is None:
            self.cb_day['values'] = []
            return
        seen = set()
        for r in self.an.rows:
            if self.an.time_idx < len(r):
                dt = parse_time(r[self.an.time_idx])
                if dt:
                    dstr = dt.date().strftime("%Y-%m-%d")
                    if dstr not in seen:
                        seen.add(dstr)
                        self.days_list.append(dstr)
        self.days_list.sort()
        self.cb_day['values'] = self.days_list
        # seleziona il primo giorno disponibile come default
        if self.days_list:
            self.sel_day.set(self.days_list[0])

    def on_select_day(self):
        """Quando si seleziona un giorno, applica automaticamente l'analisi."""
        if self.sel_day.get().strip():
            self.on_apply()

    def _current_day_index(self):
        if not self.days_list:
            return -1
        day = self.sel_day.get().strip()
        try:
            return self.days_list.index(day)
        except ValueError:
            return -1

    def on_prev_day(self):
        i = self._current_day_index()
        if i <= 0:
            return
        self.sel_day.set(self.days_list[i-1])
        self.on_select_day()

    def on_next_day(self):
        i = self._current_day_index()
        if i == -1 or i >= len(self.days_list)-1:
            return
        self.sel_day.set(self.days_list[i+1])
        self.on_select_day()

    # ---------------------- Eventi principali ----------------------
    def on_open(self):
        path = filedialog.askopenfilename(
            title="Seleziona CSV",
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        try:
            self.an = Analyzer(path)
        except Exception as e:
            messagebox.showerror("Errore", str(e))
            return
        self.current_file = path
        self.lbl_file.config(text=os.path.basename(path))
        self.populate_days()
        self.on_apply()

    def on_apply(self):
        if not self.an:
            return
        
        # Prende il giorno selezionato
        day_str = self.sel_day.get().strip()
        if not day_str:
            messagebox.showwarning("Attenzione", "Seleziona un giorno dalla lista.")
            return
        
        t_from = self._parse_date(day_str)
        if t_from is None:
            messagebox.showerror("Errore", "Formato data non valido.")
            return
        
        # Imposta t_to alla fine del giorno
        t_to = t_from.replace(hour=23, minute=59, second=59)
        
        tanks, mats, debug = self.an.analyze(
            t_from=t_from, t_to=t_to,
            include_fst=self.b_fst.get(), include_bbt=self.b_bbt.get()
        )
        self._cache_tank = tanks
        self._cache_mat = mats
        self._cache_debug = debug

        # Popola riepilogo materiali
        for r in self.tv_mat.get_children():
            self.tv_mat.delete(r)
        for m, kg, fa, n in mats:
            self.tv_mat.insert("", tk.END, values=(m, fmt_it(kg, 3), fmt_it(fa), n))

        # Popola per tank
        for r in self.tv.get_children():
            self.tv.delete(r)
        for tank, mat, g_last, v_last, sum_fa, kg_ext, n in tanks:
            self.tv.insert("", tk.END, values=(
                tank,
                mat if mat else '',
                fmt_it(g_last, 2) if g_last is not None else "",
                fmt_it(v_last, 2) if v_last is not None else "",
                fmt_it(sum_fa),
                fmt_it(kg_ext, 3),
                n
            ))

        # Popola debug
        for r in self.tv_debug.get_children():
            self.tv_debug.delete(r)
        debug_total = 0.0
        for dt, tank, mat, g, v, fa, kg in debug:
            self.tv_debug.insert("", tk.END, values=(
                dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                tank,
                mat,
                fmt_it(g, 2) if g is not None else "",
                fmt_it(v, 2) if v is not None else "",
                fmt_it(fa, 6) if fa is not None else "",
                fmt_it(kg, 3)
            ))
            debug_total += kg
        
        self.lbl_debug_total.config(text=f"Totale Kg estratto (debug): {fmt_it(debug_total, 3)} | Righe: {len(debug)}")

        # Aggiorna eventuale finestra totale aperta
        self.refresh_total_window()
        
        # Aggiorna tabella variazioni
        self.update_variations_table()

    def update_variations_table(self):
        """Calcola e mostra le variazioni di livello e Kg tra misurazioni successive"""
        if not self._cache_debug:
            return
        
        # Pulisci tabella
        for r in self.tv_variations.get_children():
            self.tv_variations.delete(r)
        
        # Raggruppa per tank
        by_tank = defaultdict(list)
        for dt, tank, mat, g, v, fa, kg in self._cache_debug:
            by_tank[tank].append((dt, mat, g, v, fa, kg))
        
        # Ordina per timestamp
        for tank in by_tank:
            by_tank[tank].sort(key=lambda x: x[0] if x[0] else datetime.min)
        
        # Aggiorna lista tank disponibili
        all_tanks = sorted(by_tank.keys())
        self.cb_filter_tank['values'] = ["Tutti"] + all_tanks
        
        # Filtra per tank selezionato
        filter_tank = self.var_filter_tank.get()
        if filter_tank != "Tutti" and filter_tank in by_tank:
            tanks_to_show = {filter_tank: by_tank[filter_tank]}
        else:
            tanks_to_show = by_tank
        
        # Calcola variazioni
        variations = []
        total_delta_level = 0.0
        total_delta_kg = 0.0
        max_increase_level = 0.0
        max_decrease_level = 0.0
        max_increase_kg = 0.0
        max_decrease_kg = 0.0
        
        for tank, measurements in tanks_to_show.items():
            for i in range(1, len(measurements)):
                prev = measurements[i-1]
                curr = measurements[i]
                
                dt_prev, mat_prev, g_prev, v_prev, fa_prev, kg_prev = prev
                dt_curr, mat_curr, g_curr, v_curr, fa_curr, kg_curr = curr
                
                delta_level = v_curr - v_prev if (v_curr is not None and v_prev is not None) else None
                delta_kg = kg_curr - kg_prev
                
                if delta_level is not None:
                    total_delta_level += delta_level
                    if delta_level > max_increase_level:
                        max_increase_level = delta_level
                    if delta_level < max_decrease_level:
                        max_decrease_level = delta_level
                
                total_delta_kg += delta_kg
                if delta_kg > max_increase_kg:
                    max_increase_kg = delta_kg
                if delta_kg < max_decrease_kg:
                    max_decrease_kg = delta_kg
                
                # Colora le righe in base alla variazione
                tag = ""
                if delta_level is not None:
                    if delta_level < -10:  # Calo significativo
                        tag = "decrease"
                    elif delta_level > 10:  # Aumento significativo
                        tag = "increase"
                
                item = self.tv_variations.insert("", tk.END, values=(
                    dt_curr.strftime("%Y-%m-%d %H:%M:%S") if dt_curr else "",
                    tank,
                    mat_curr,
                    fmt_it(v_prev, 2) if v_prev is not None else "",
                    fmt_it(v_curr, 2) if v_curr is not None else "",
                    fmt_it(delta_level, 2) if delta_level is not None else "",
                    fmt_it(g_prev, 2) if g_prev is not None else "",
                    fmt_it(g_curr, 2) if g_curr is not None else "",
                    fmt_it(kg_prev, 3),
                    fmt_it(kg_curr, 3),
                    fmt_it(delta_kg, 3)
                ), tags=(tag,))
                
                variations.append((dt_curr, tank, mat_curr, v_prev, v_curr, delta_level, g_prev, g_curr, kg_prev, kg_curr, delta_kg))
        
        # Configura colori
        self.tv_variations.tag_configure("decrease", background="#ffcccc")  # Rosso chiaro
        self.tv_variations.tag_configure("increase", background="#ccffcc")  # Verde chiaro
        
        # Aggiorna summary
        num_variations = len(variations)
        if num_variations > 0:
            summary_text = (
                f"Variazioni totali: {num_variations} | "
                f"ΔLevel totale: {fmt_it(total_delta_level, 2)} hl | "
                f"ΔKg totale: {fmt_it(total_delta_kg, 3)} kg\n"
                f"Max aumento level: {fmt_it(max_increase_level, 2)} hl | "
                f"Max calo level: {fmt_it(max_decrease_level, 2)} hl | "
                f"Max aumento Kg: {fmt_it(max_increase_kg, 3)} kg | "
                f"Max perdita Kg: {fmt_it(max_decrease_kg, 3)} kg"
            )
        else:
            summary_text = "Nessuna variazione disponibile"
        
        self.lbl_var_summary.config(text=summary_text)

    def _parse_date(self, s):
        s = str(s).strip()
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None

    # ---------------------- Grafici ----------------------
    def on_generate_charts(self):
        if not self.an or not _HAS_MATPLOTLIB:
            return
        
        # Cancella grafici precedenti
        for widget in self.charts_frame.winfo_children():
            widget.destroy()
        
        daily_data = self.an.analyze_all_days(
            include_fst=self.b_fst.get(),
            include_bbt=self.b_bbt.get()
        )
        
        if not daily_data:
            ttk.Label(self.charts_frame, text="Nessun dato disponibile per i grafici").pack(pady=20)
            return
        
        # Prepara dati
        days = sorted(daily_data.keys())
        kg_totals = [daily_data[d]['kg'] for d in days]
        
        # Grafico 1: Totale Kg estratto per giorno
        fig1 = Figure(figsize=(12, 4), dpi=80)
        ax1 = fig1.add_subplot(111)
        ax1.plot(days, kg_totals, marker='o', linewidth=2, markersize=6)
        ax1.set_title('Kg Estratto Totale per Giorno')
        ax1.set_xlabel('Giorno')
        ax1.set_ylabel('Kg Estratto')
        ax1.grid(True, alpha=0.3)
        ax1.tick_params(axis='x', rotation=45)
        fig1.tight_layout()
        
        canvas1 = FigureCanvasTkAgg(fig1, self.charts_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Grafico 2: Per materiale (top 5)
        all_materials = set()
        for d in days:
            all_materials.update(daily_data[d]['by_material'].keys())
        
        # Calcola totali per materiale
        material_totals = {}
        for mat in all_materials:
            material_totals[mat] = sum(daily_data[d]['by_material'].get(mat, 0) for d in days)
        
        # Prendi top 5
        top_materials = sorted(material_totals.items(), key=lambda x: -x[1])[:5]
        
        if top_materials:
            fig2 = Figure(figsize=(12, 4), dpi=80)
            ax2 = fig2.add_subplot(111)
            
            for mat, _ in top_materials:
                values = [daily_data[d]['by_material'].get(mat, 0) for d in days]
                ax2.plot(days, values, marker='o', label=mat, linewidth=2, markersize=5)
            
            ax2.set_title('Kg Estratto per Materiale (Top 5)')
            ax2.set_xlabel('Giorno')
            ax2.set_ylabel('Kg Estratto')
            ax2.legend(loc='best')
            ax2.grid(True, alpha=0.3)
            ax2.tick_params(axis='x', rotation=45)
            fig2.tight_layout()
            
            canvas2 = FigureCanvasTkAgg(fig2, self.charts_frame)
            canvas2.draw()
            canvas2.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Grafico 3: Per tank (top 5)
        all_tanks = set()
        for d in days:
            all_tanks.update(daily_data[d]['by_tank'].keys())
        
        tank_totals = {}
        for tank in all_tanks:
            tank_totals[tank] = sum(daily_data[d]['by_tank'].get(tank, 0) for d in days)
        
        top_tanks = sorted(tank_totals.items(), key=lambda x: -x[1])[:5]
        
        if top_tanks:
            fig3 = Figure(figsize=(12, 4), dpi=80)
            ax3 = fig3.add_subplot(111)
            
            for tank, _ in top_tanks:
                values = [daily_data[d]['by_tank'].get(tank, 0) for d in days]
                ax3.plot(days, values, marker='o', label=tank, linewidth=2, markersize=5)
            
            ax3.set_title('Kg Estratto per Tank (Top 5)')
            ax3.set_xlabel('Giorno')
            ax3.set_ylabel('Kg Estratto')
            ax3.legend(loc='best')
            ax3.grid(True, alpha=0.3)
            ax3.tick_params(axis='x', rotation=45)
            fig3.tight_layout()
            
            canvas3 = FigureCanvasTkAgg(fig3, self.charts_frame)
            canvas3.draw()
            canvas3.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=5)

    # ---------------------- Totale Cantina ----------------------
    def compute_totals(self):
        if not self._cache_mat:
            return 0.0, 0.0, 0
        excl0 = self.var_exclude_mat0.get()
        tot_kg = 0.0
        tot_fa = 0.0
        tot_n  = 0
        for m, kg, fa, n in self._cache_mat:
            if excl0 and str(m).strip() == '0':
                continue
            tot_kg += kg
            tot_fa += fa
            tot_n  += n
        return tot_kg, tot_fa, tot_n

    def on_show_total(self):
        if self._tot_win and tk.Toplevel.winfo_exists(self._tot_win):
            self._tot_win.lift()
            self.refresh_total_window()
            return
        self._tot_win = tk.Toplevel(self)
        self._tot_win.title("Totale Cantina")
        self._tot_win.resizable(False, False)
        frm = ttk.Frame(self._tot_win, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)
        self.lbl_tot_kg = ttk.Label(frm, text="Totale Kg estratto: -", font=("Segoe UI", 10, "bold"))
        self.lbl_tot_fa = ttk.Label(frm, text="Totale f(A): -")
        self.lbl_tot_n  = ttk.Label(frm, text="Misure conteggiate: -")
        self.lbl_note   = ttk.Label(frm, text="", foreground="#555")
        self.lbl_tot_kg.pack(anchor=tk.W)
        self.lbl_tot_fa.pack(anchor=tk.W, pady=(4,0))
        self.lbl_tot_n.pack(anchor=tk.W, pady=(4,6))
        self.lbl_note.pack(anchor=tk.W)
        btns = ttk.Frame(frm)
        btns.pack(fill=tk.X, pady=(10,0))
        ttk.Button(btns, text="Chiudi", command=self._tot_win.destroy).pack(side=tk.RIGHT)
        self.refresh_total_window()

    def refresh_total_window(self):
        if not (self._tot_win and tk.Toplevel.winfo_exists(self._tot_win)):
            return
        kg, fa, n = self.compute_totals()
        self.lbl_tot_kg.config(text=f"Totale Kg estratto: {fmt_it(kg, 3)}")
        self.lbl_tot_fa.config(text=f"Totale f(A): {fmt_it(fa)}")
        self.lbl_tot_n.config(text=f"Misure conteggiate: {n}")
        note = "(Material=0 escluso)" if self.var_exclude_mat0.get() else "(Material=0 incluso)"
        self.lbl_note.config(text=note)

    # ---------------------- Export ----------------------
    def on_export_tank_csv(self):
        if not self._cache_tank:
            return
        path = filedialog.asksaveasfilename(
            title="Salva CSV (per Tank)",
            defaultextension=".csv",
            initialfile="per_tank_gravity_volume_material_fa_kg.csv",
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Tank','Materiale','Gravity_ultimo','Volume_ultimo','Somma_f(A)','Kg_estratto','Misure'])
                for tank, mat, g_last, v_last, sum_fa, kg_ext, n in self._cache_tank:
                    w.writerow([
                        tank,
                        mat or '',
                        f"{g_last:.2f}" if g_last is not None else "",
                        f"{v_last:.2f}" if v_last is not None else "",
                        f"{sum_fa:.6f}",
                        f"{kg_ext:.3f}",
                        n
                    ])
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def on_export_mat_csv(self):
        if not self._cache_mat:
            return
        path = filedialog.asksaveasfilename(
            title="Salva CSV (per Materiale)",
            defaultextension=".csv",
            initialfile="per_material_somma_kg_fa.csv",
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Materiale','Kg_estratto','Somma_f(A)','Misure'])
                for m, kg, fa, n in self._cache_mat:
                    w.writerow([m, f"{kg:.3f}", f"{fa:.6f}", n])
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def on_export_debug_csv(self):
        if not self._cache_debug:
            return
        path = filedialog.asksaveasfilename(
            title="Salva CSV Debug",
            defaultextension=".csv",
            initialfile="debug_dettaglio_calcoli.csv",
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Timestamp','Tank','Materiale','Gravity','Level_hl','f(A)','Kg_estratto'])
                for dt, tank, mat, g, v, fa, kg in self._cache_debug:
                    w.writerow([
                        dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                        tank,
                        mat,
                        f"{g:.2f}" if g is not None else "",
                        f"{v:.2f}" if v is not None else "",
                        f"{fa:.6f}" if fa is not None else "",
                        f"{kg:.3f}"
                    ])
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def on_export_variations_csv(self):
        """Esporta le variazioni in CSV"""
        if not self._cache_debug:
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva CSV Variazioni",
            defaultextension=".csv",
            initialfile="variazioni_level_kg.csv",
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            # Calcola variazioni
            by_tank = defaultdict(list)
            for dt, tank, mat, g, v, fa, kg in self._cache_debug:
                by_tank[tank].append((dt, mat, g, v, fa, kg))
            
            for tank in by_tank:
                by_tank[tank].sort(key=lambda x: x[0] if x[0] else datetime.min)
            
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Timestamp','Tank','Materiale','Level_Prec_hl','Level_Corr_hl','Delta_Level_hl',
                           'Gravity_Prec','Gravity_Corr','Kg_Prec','Kg_Corr','Delta_Kg'])
                
                for tank, measurements in sorted(by_tank.items()):
                    for i in range(1, len(measurements)):
                        prev = measurements[i-1]
                        curr = measurements[i]
                        
                        dt_prev, mat_prev, g_prev, v_prev, fa_prev, kg_prev = prev
                        dt_curr, mat_curr, g_curr, v_curr, fa_curr, kg_curr = curr
                        
                        delta_level = v_curr - v_prev if (v_curr is not None and v_prev is not None) else None
                        delta_kg = kg_curr - kg_prev
                        
                        w.writerow([
                            dt_curr.strftime("%Y-%m-%d %H:%M:%S") if dt_curr else "",
                            tank,
                            mat_curr,
                            f"{v_prev:.2f}" if v_prev is not None else "",
                            f"{v_curr:.2f}" if v_curr is not None else "",
                            f"{delta_level:.2f}" if delta_level is not None else "",
                            f"{g_prev:.2f}" if g_prev is not None else "",
                            f"{g_curr:.2f}" if g_curr is not None else "",
                            f"{kg_prev:.3f}",
                            f"{kg_curr:.3f}",
                            f"{delta_kg:.3f}"
                        ])
            
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def on_export_xlsx(self):
        if not self._cache_mat and not self._cache_tank:
            return
        if not _HAS_OPENPYXL:
            return
        path = filedialog.asksaveasfilename(
            title="Salva XLSX",
            defaultextension=".xlsx",
            initialfile="report_tank_material_fa_kg.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'Per Materiale'
            ws1.append(['Materiale','Kg_estratto','Somma_f(A)','Misure'])
            for m, kg, fa, n in self._cache_mat:
                ws1.append([m, float(f"{kg:.3f}"), float(f"{fa:.6f}"), n])

            ws2 = wb.create_sheet('Per Tank')
            ws2.append(['Tank','Materiale','Gravity_ultimo','Volume_ultimo','Somma_f(A)','Kg_estratto','Misure'])
            for tank, mat, g_last, v_last, sum_fa, kg_ext, n in self._cache_tank:
                ws2.append([
                    tank,
                    mat or '',
                    float(f"{g_last:.2f}") if g_last is not None else None,
                    float(f"{v_last:.2f}") if v_last is not None else None,
                    float(f"{sum_fa:.6f}"),
                    float(f"{kg_ext:.3f}"),
                    n
                ])

            ws3 = wb.create_sheet('Debug')
            ws3.append(['Timestamp','Tank','Materiale','Gravity','Level_hl','f(A)','Kg_estratto'])
            for dt, tank, mat, g, v, fa, kg in self._cache_debug:
                ws3.append([
                    dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                    tank,
                    mat,
                    float(f"{g:.2f}") if g is not None else None,
                    float(f"{v:.2f}") if v is not None else None,
                    float(f"{fa:.6f}") if fa is not None else None,
                    float(f"{kg:.3f}")
                ])

            # Foglio Variazioni
            ws_var = wb.create_sheet('Variazioni')
            ws_var.append(['Timestamp','Tank','Materiale','Level_Prec_hl','Level_Corr_hl','Delta_Level_hl',
                          'Gravity_Prec','Gravity_Corr','Kg_Prec','Kg_Corr','Delta_Kg'])
            
            by_tank = defaultdict(list)
            for dt, tank, mat, g, v, fa, kg in self._cache_debug:
                by_tank[tank].append((dt, mat, g, v, fa, kg))
            
            for tank in by_tank:
                by_tank[tank].sort(key=lambda x: x[0] if x[0] else datetime.min)
            
            for tank, measurements in sorted(by_tank.items()):
                for i in range(1, len(measurements)):
                    prev = measurements[i-1]
                    curr = measurements[i]
                    
                    dt_prev, mat_prev, g_prev, v_prev, fa_prev, kg_prev = prev
                    dt_curr, mat_curr, g_curr, v_curr, fa_curr, kg_curr = curr
                    
                    delta_level = v_curr - v_prev if (v_curr is not None and v_prev is not None) else None
                    delta_kg = kg_curr - kg_prev
                    
                    ws_var.append([
                        dt_curr.strftime("%Y-%m-%d %H:%M:%S") if dt_curr else "",
                        tank,
                        mat_curr,
                        float(f"{v_prev:.2f}") if v_prev is not None else None,
                        float(f"{v_curr:.2f}") if v_curr is not None else None,
                        float(f"{delta_level:.2f}") if delta_level is not None else None,
                        float(f"{g_prev:.2f}") if g_prev is not None else None,
                        float(f"{g_curr:.2f}") if g_curr is not None else None,
                        float(f"{kg_prev:.3f}"),
                        float(f"{kg_curr:.3f}"),
                        float(f"{delta_kg:.3f}")
                    ])

            ws4 = wb.create_sheet('Note')
            ws4.append(['Descrizione','Valore'])
            ws4.append(['Equivalenza', "'Average Gravity' == 'Average Plato' (usati come 'Gravity')"])
            ws4.append(['f(A)', '((0.0000188792*G + 0.003646886)*G + 1.001077)*G - 0.01223565'])
            ws4.append(['Kg estratto (riga)', 'f(A) * Level'])
            ws4.append(['Aggregazioni', 'Somme su periodo filtrato; Material per riga secondo colonna Material del tank'])
            ws4.append(['Mapping Material', '7=ichnusa; 8=non filtrata; 9=cruda; 28=ambra limpida'])
            excl = 'sì' if self.var_exclude_mat0.get() else 'no'
            ws4.append(['Totale Cantina', f"Material=0 escluso: {excl}"])
            ws4.append(['Modalità', 'Giorno singolo'])
            ws4.append(['', ''])
            ws4.append(['Tool Info', ''])
            ws4.append(['Sviluppato da', APP_AUTHOR])
            ws4.append(['Dipartimento', APP_DEPT])
            ws4.append(['Versione', APP_VERSION])
            ws4.append(['Contatto', APP_EMAIL])

            wb.save(path)
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))


if __name__ == '__main__':
    app = App()
    app.mainloop()