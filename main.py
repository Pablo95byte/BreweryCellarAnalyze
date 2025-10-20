#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tank Analysis Tool - Applicazione GUI
Sviluppato da PA (ASS_ST)
"""

import os
import csv
from datetime import datetime
from collections import defaultdict

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Import moduli custom
from config import (
    APP_TITLE, APP_VERSION, APP_AUTHOR, APP_EMAIL, APP_DEPT,
    WINDOW_SIZE, WINDOW_MIN_SIZE, SPLASH_DURATION,
    COLORS, THRESHOLDS, EXPORT_FILENAMES, FA_FORMULA
)
from utils import fmt_it, parse_time
from analyzer import TankAnalyzer

# Import opzionali
try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.backends.backend_pdf import PdfPages
    HAS_MATPLOTLIB = True
except Exception:
    HAS_MATPLOTLIB = False


class TankAnalysisApp(tk.Tk):
    """Applicazione principale per l'analisi dei tank"""
    
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry(WINDOW_SIZE)
        self.minsize(*WINDOW_MIN_SIZE)
        
        # State
        self.analyzer = None
        self.current_file = None
        self.days_list = []
        self.sel_day = tk.StringVar(value="")
        
        # Cache risultati
        self._cache_tank = []
        self._cache_mat = []
        self._cache_debug = []
        
        # Variabili UI
        self.var_exclude_mat0 = tk.BooleanVar(value=False)
        self.var_filter_tank = tk.StringVar(value="Tutti")
        self.b_fst = tk.BooleanVar(value=True)
        self.b_bbt = tk.BooleanVar(value=True)
        self.b_rbt = tk.BooleanVar(value=True)
        
        # Windows
        self._tot_win = None
        
        # Grafici
        self.chart_figures = []
        
        # Mostra splash e costruisci UI
        self.show_splash()
        self._build_ui()
        
        # Inizializza cache variazioni
        self._cache_variations = []
    
    # ==================== SPLASH SCREEN ====================
    
    def show_splash(self):
        """Mostra splash screen all'avvio"""
        splash = tk.Toplevel(self)
        splash.title("")
        splash.overrideredirect(True)
        
        w, h = 400, 250
        x = (splash.winfo_screenwidth() // 2) - (w // 2)
        y = (splash.winfo_screenheight() // 2) - (h // 2)
        splash.geometry(f"{w}x{h}+{x}+{y}")
        
        frame = ttk.Frame(splash, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=APP_TITLE, font=("Segoe UI", 16, "bold")).pack(pady=(20,10))
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        ttk.Label(frame, text=f"Versione {APP_VERSION}", font=("Segoe UI", 11)).pack(pady=5)
        ttk.Label(frame, text=f"Sviluppato da {APP_AUTHOR}", font=("Segoe UI", 10)).pack(pady=5)
        ttk.Label(frame, text=APP_DEPT, font=("Segoe UI", 9), foreground="#666").pack(pady=2)
        ttk.Label(frame, text=APP_EMAIL, font=("Segoe UI", 9), foreground=COLORS['link']).pack(pady=2)
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        ttk.Label(frame, text="Caricamento in corso...", font=("Segoe UI", 9), foreground="#999").pack(pady=(10,20))
        
        splash.update()
        self.after(SPLASH_DURATION, splash.destroy)
    
    # ==================== BUILD UI ====================
    
    def _build_ui(self):
        """Costruisce l'interfaccia utente"""
        self._build_menubar()
        self._build_toolbar()
        self._build_filters()
        self._build_day_selector()
        self._build_notebook()
        self._build_actions()
        self._build_footer()
    
    def _build_menubar(self):
        """Crea la barra menu"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        menubar.add_command(label="Formule", command=self.show_formulas)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
    
    def _build_toolbar(self):
        """Toolbar superiore"""
        top = ttk.Frame(self)
        top.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(top, text="Apri CSV...", command=self.on_open).pack(side=tk.LEFT)
        self.lbl_file = ttk.Label(top, text="Nessun file caricato", foreground=COLORS['info'])
        self.lbl_file.pack(side=tk.LEFT, padx=10)
    
    def _build_filters(self):
        """Filtri FST/BBT/RBT"""
        filt = ttk.Frame(self)
        filt.pack(fill=tk.X, padx=10, pady=(0,8))
        
        ttk.Checkbutton(filt, text="FST", variable=self.b_fst).pack(side=tk.LEFT, padx=(0,10))
        ttk.Checkbutton(filt, text="BBT", variable=self.b_bbt).pack(side=tk.LEFT, padx=(0,10))
        ttk.Checkbutton(filt, text="RBT", variable=self.b_rbt).pack(side=tk.LEFT, padx=(0,10))
        ttk.Button(filt, text="Applica", command=self.on_apply).pack(side=tk.LEFT)
    
    def _build_day_selector(self):
        """Selettore giorno"""
        frame = ttk.Frame(self)
        frame.pack(fill=tk.X, padx=10, pady=(0,8))
        
        ttk.Button(frame, text="◀", width=2, command=self.on_prev_day).pack(side=tk.LEFT, padx=(0,5))
        ttk.Label(frame, text="Giorno:").pack(side=tk.LEFT)
        self.cb_day = ttk.Combobox(frame, textvariable=self.sel_day, width=12, state="readonly")
        self.cb_day.pack(side=tk.LEFT, padx=(5,5))
        self.cb_day.bind("<<ComboboxSelected>>", lambda e: self.on_select_day())
        ttk.Button(frame, text="▶", width=2, command=self.on_next_day).pack(side=tk.LEFT)
    
    def _build_notebook(self):
        """Crea il notebook con tutti i tab"""
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        # Tab sempre presenti
        self._build_summary_tab()
        self._build_debug_tab()
        
        # Tab grafici (se matplotlib disponibile)
        if HAS_MATPLOTLIB:
            self._build_charts_tab()
        else:
            # Mostra tab placeholder se matplotlib non disponibile
            self._build_charts_placeholder_tab()
        
        self._build_variations_tab()
        self._build_raw_data_tab()
    
    def _build_actions(self):
        """Bottoni azioni"""
        actions = ttk.Frame(self)
        actions.pack(fill=tk.X, padx=10, pady=(0,10))
        
        ttk.Button(actions, text="Esporta per Tank (CSV)", command=self.on_export_tank_csv).pack(side=tk.LEFT)
        ttk.Button(actions, text="Esporta per Materiale (CSV)", command=self.on_export_mat_csv).pack(side=tk.LEFT, padx=(10,0))
        ttk.Button(actions, text="Esporta Debug (CSV)", command=self.on_export_debug_csv).pack(side=tk.LEFT, padx=(10,0))
        ttk.Button(actions, text="Esporta Variazioni (CSV)", command=self.on_export_variations_csv).pack(side=tk.LEFT, padx=(10,0))
        
        btnx = ttk.Button(actions, text="Esporta report (XLSX)", command=self.on_export_xlsx)
        if not HAS_OPENPYXL:
            btnx.state(["disabled"])
        btnx.pack(side=tk.LEFT, padx=(10,0))
    
    def _build_footer(self):
        """Footer con credits"""
        hint = ttk.Label(self, text=(
            "Average Gravity == Average Plato (valore usato come Gravity). "
            f"{FA_FORMULA}; "
            "Kg estratto (riga) = f(A) * Level. Mapping: 7=ichnusa, 8=non filtrata, 9=cruda, 28=ambra limpida."
        ), foreground=COLORS['info'])
        hint.pack(fill=tk.X, padx=10, pady=(0,5))
        
        footer = ttk.Frame(self)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer_text = f"Sviluppato da {APP_AUTHOR} ({APP_DEPT}) - v{APP_VERSION} - {APP_EMAIL}"
        ttk.Label(footer, text=footer_text, foreground="#888", font=("Segoe UI", 8)).pack(pady=5)
    
    # ==================== TAB: RIEPILOGO ====================
    
    def _build_summary_tab(self):
        """Tab riepilogo"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Riepilogo")
        
        # Riepilogo per materiale
        mat_frame = ttk.LabelFrame(tab, text="Riepilogo per Materiale")
        mat_frame.pack(fill=tk.X, padx=10, pady=10)
        
        cols_m = ("m", "kg", "fa", "n")
        self.tv_mat = ttk.Treeview(mat_frame, columns=cols_m, show="headings", height=6)
        self.tv_mat.heading("m", text="Materiale")
        self.tv_mat.heading("kg", text="Kg estratto (somma)")
        self.tv_mat.heading("fa", text="Somma f(A)")
        self.tv_mat.heading("n", text="N. Tank")
        self.tv_mat.column("m", width=240, anchor=tk.W)
        self.tv_mat.column("kg", width=160, anchor=tk.E)
        self.tv_mat.column("fa", width=160, anchor=tk.E)
        self.tv_mat.column("n", width=100, anchor=tk.E)
        self.tv_mat.pack(fill=tk.X, padx=6, pady=6)
        
        # Barra totale
        tot_bar = ttk.Frame(tab)
        tot_bar.pack(fill=tk.X, padx=10, pady=(0,8))
        ttk.Checkbutton(tot_bar, text="Escludi Material = 0 dal totale", 
                       variable=self.var_exclude_mat0, command=self.refresh_total_window).pack(side=tk.LEFT)
        ttk.Button(tot_bar, text="Totale Cantina...", command=self.on_show_total).pack(side=tk.LEFT, padx=(10,0))
        
        # Tabella per tank
        tank_frame = ttk.LabelFrame(tab, text="Per Tank")
        tank_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        cols = ("tank", "mat", "g_last", "v_last", "sum_fa", "kg_ext", "n")
        self.tv = ttk.Treeview(tank_frame, columns=cols, show="headings", height=16)
        self.tv.heading("tank", text="Tank")
        self.tv.heading("mat", text="Materiale")
        self.tv.heading("g_last", text="Gravity (plato)")
        self.tv.heading("v_last", text="Volume (hl)")
        self.tv.heading("sum_fa", text="Somma f(A)")
        self.tv.heading("kg_ext", text="Kg estratto")
        self.tv.heading("n", text="Misure")
        self.tv.column("tank", width=100, anchor=tk.W)
        self.tv.column("mat", width=200, anchor=tk.W)
        self.tv.column("g_last", width=140, anchor=tk.E)
        self.tv.column("v_last", width=140, anchor=tk.E)
        self.tv.column("sum_fa", width=160, anchor=tk.E)
        self.tv.column("kg_ext", width=140, anchor=tk.E)
        self.tv.column("n", width=80, anchor=tk.E)
        
        vsb = ttk.Scrollbar(tank_frame, orient="vertical", command=self.tv.yview)
        self.tv.configure(yscrollcommand=vsb.set)
        self.tv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.LEFT, fill=tk.Y)
    
    # ==================== TAB: DEBUG ====================
    
    def _build_debug_tab(self):
        """Tab debug dati"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Debug Dati")
        
        ttk.Label(tab, text="Dettaglio di ogni riga elaborata nel giorno selezionato", 
                 foreground=COLORS['info']).pack(fill=tk.X, padx=10, pady=10)
        
        debug_frame = ttk.Frame(tab)
        debug_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        cols_d = ("time", "tank", "material", "gravity", "level", "fA", "kg")
        self.tv_debug = ttk.Treeview(debug_frame, columns=cols_d, show="headings", height=20)
        self.tv_debug.heading("time", text="Timestamp")
        self.tv_debug.heading("tank", text="Tank")
        self.tv_debug.heading("material", text="Materiale")
        self.tv_debug.heading("gravity", text="Gravity")
        self.tv_debug.heading("level", text="Level (hl)")
        self.tv_debug.heading("fA", text="f(A)")
        self.tv_debug.heading("kg", text="Kg estratto")
        
        self.tv_debug.column("time", width=150, anchor=tk.W)
        self.tv_debug.column("tank", width=80, anchor=tk.W)
        self.tv_debug.column("material", width=150, anchor=tk.W)
        self.tv_debug.column("gravity", width=100, anchor=tk.E)
        self.tv_debug.column("level", width=100, anchor=tk.E)
        self.tv_debug.column("fA", width=120, anchor=tk.E)
        self.tv_debug.column("kg", width=120, anchor=tk.E)
        
        vsb_d = ttk.Scrollbar(debug_frame, orient="vertical", command=self.tv_debug.yview)
        self.tv_debug.configure(yscrollcommand=vsb_d.set)
        self.tv_debug.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb_d.pack(side=tk.LEFT, fill=tk.Y)
        
        self.lbl_debug_total = ttk.Label(tab, text="Totale Kg estratto (debug): -", font=("Segoe UI", 10, "bold"))
        self.lbl_debug_total.pack(fill=tk.X, padx=10, pady=(5,10))
    
    # ==================== TAB: GRAFICI ====================
    
    def _build_charts_placeholder_tab(self):
        """Tab placeholder se matplotlib non disponibile"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Grafici")
        
        msg_frame = ttk.Frame(tab)
        msg_frame.pack(expand=True)
        
        ttk.Label(msg_frame, text="⚠️ Matplotlib non disponibile", 
                 font=("Segoe UI", 12, "bold")).pack(pady=10)
        ttk.Label(msg_frame, text="Per utilizzare i grafici, installa matplotlib:", 
                 foreground=COLORS['info']).pack(pady=5)
        ttk.Label(msg_frame, text="pip install matplotlib", 
                 font=("Courier New", 10), foreground=COLORS['link']).pack(pady=5)
    
    def _build_charts_tab(self):
        """Tab grafici"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Grafici")
        
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(btn_frame, text="Genera Grafici", command=self.on_generate_charts).pack(side=tk.LEFT)
        ttk.Label(btn_frame, text="(analizza tutti i giorni disponibili)", foreground=COLORS['info']).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Esporta Grafici (PDF)", command=self.on_export_charts_pdf).pack(side=tk.LEFT, padx=(20,0))
        
        self.charts_frame = ttk.Frame(tab)
        self.charts_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
    
    # ==================== TAB: VARIAZIONI ====================
    
    def _build_variations_tab(self):
        """Tab analisi variazioni"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Analisi Variazioni")
        
        ttk.Label(tab, text="Variazioni giornaliere di livello e Kg estratto (confronto ultimo valore di ogni giorno)", 
                 foreground=COLORS['info']).pack(fill=tk.X, padx=10, pady=10)
        
        # Filtro tank
        select_frame = ttk.Frame(tab)
        select_frame.pack(fill=tk.X, padx=10, pady=(0,10))
        ttk.Label(select_frame, text="Filtra per Tank:").pack(side=tk.LEFT)
        self.cb_filter_tank = ttk.Combobox(select_frame, textvariable=self.var_filter_tank, width=15, state="readonly")
        self.cb_filter_tank['values'] = ["Tutti"]
        self.cb_filter_tank.pack(side=tk.LEFT, padx=(5,10))
        self.cb_filter_tank.bind("<<ComboboxSelected>>", lambda e: self.update_variations_table())
        ttk.Button(select_frame, text="Aggiorna Filtro", command=lambda: self.update_variations_table()).pack(side=tk.LEFT)
        ttk.Button(select_frame, text="Carica Variazioni", command=lambda: self.load_all_variations()).pack(side=tk.LEFT, padx=(10,0))
        
        # Tabella variazioni
        var_frame = ttk.Frame(tab)
        var_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        cols_v = ("date", "tank", "material", "level_prev", "level_curr", "delta_level", 
                 "gravity_prev", "gravity_curr", "kg_prev", "kg_curr", "delta_kg")
        self.tv_variations = ttk.Treeview(var_frame, columns=cols_v, show="headings", height=20)
        self.tv_variations.heading("date", text="Data")
        self.tv_variations.heading("tank", text="Tank")
        self.tv_variations.heading("material", text="Materiale")
        self.tv_variations.heading("level_prev", text="Level Prec (hl)")
        self.tv_variations.heading("level_curr", text="Level Corr (hl)")
        self.tv_variations.heading("delta_level", text="ΔLevel (hl)")
        self.tv_variations.heading("gravity_prev", text="Gravity Prec")
        self.tv_variations.heading("gravity_curr", text="Gravity Corr")
        self.tv_variations.heading("kg_prev", text="Kg Prec")
        self.tv_variations.heading("kg_curr", text="Kg Corr")
        self.tv_variations.heading("delta_kg", text="ΔKg")
        
        self.tv_variations.column("date", width=100, anchor=tk.W)
        self.tv_variations.column("tank", width=70, anchor=tk.W)
        self.tv_variations.column("material", width=120, anchor=tk.W)
        for col in ["level_prev", "level_curr", "delta_level", "gravity_prev", "gravity_curr", "kg_prev", "kg_curr", "delta_kg"]:
            self.tv_variations.column(col, width=90, anchor=tk.E)
        
        vsb_v = ttk.Scrollbar(var_frame, orient="vertical", command=self.tv_variations.yview)
        hsb_v = ttk.Scrollbar(var_frame, orient="horizontal", command=self.tv_variations.xview)
        self.tv_variations.configure(yscrollcommand=vsb_v.set, xscrollcommand=hsb_v.set)
        self.tv_variations.grid(row=0, column=0, sticky='nsew')
        vsb_v.grid(row=0, column=1, sticky='ns')
        hsb_v.grid(row=1, column=0, sticky='ew')
        var_frame.grid_rowconfigure(0, weight=1)
        var_frame.grid_columnconfigure(0, weight=1)
        
        # Summary
        summary_frame = ttk.LabelFrame(tab, text="Riepilogo Variazioni")
        summary_frame.pack(fill=tk.X, padx=10, pady=(0,10))
        self.lbl_var_summary = ttk.Label(summary_frame, text="Clicca 'Carica Variazioni' per analizzare tutti i giorni del CSV", 
                                         foreground=COLORS['info'])
        self.lbl_var_summary.pack(padx=10, pady=10, anchor=tk.W)
        
        # Configura tag colori per la treeview variazioni
        self.tv_variations.tag_configure("decrease", background=COLORS['decrease'])
        self.tv_variations.tag_configure("increase", background=COLORS['increase'])
    
    # ==================== TAB: DATI RAW ====================
    
    def _build_raw_data_tab(self):
        """Tab dati raw"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Dati Raw")
        
        info_frame = ttk.Frame(tab)
        info_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Label(info_frame, text="Visualizzazione di TUTTI i dati grezzi del CSV caricato (nessun filtro applicato)", 
                 foreground=COLORS['info']).pack(side=tk.LEFT)
        ttk.Button(info_frame, text="Esporta Raw (CSV)", command=self.on_export_raw_csv).pack(side=tk.RIGHT, padx=(10,0))
        ttk.Button(info_frame, text="Aggiorna", command=self.populate_raw_data).pack(side=tk.RIGHT)
        
        raw_frame = ttk.Frame(tab)
        raw_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        
        self.tv_raw = ttk.Treeview(raw_frame, show="headings", height=20)
        
        vsb_raw = ttk.Scrollbar(raw_frame, orient="vertical", command=self.tv_raw.yview)
        hsb_raw = ttk.Scrollbar(raw_frame, orient="horizontal", command=self.tv_raw.xview)
        self.tv_raw.configure(yscrollcommand=vsb_raw.set, xscrollcommand=hsb_raw.set)
        
        self.tv_raw.grid(row=0, column=0, sticky='nsew')
        vsb_raw.grid(row=0, column=1, sticky='ns')
        hsb_raw.grid(row=1, column=0, sticky='ew')
        raw_frame.grid_rowconfigure(0, weight=1)
        raw_frame.grid_columnconfigure(0, weight=1)
        
        self.lbl_raw_info = ttk.Label(tab, text="Nessun file caricato", foreground=COLORS['info'])
        self.lbl_raw_info.pack(fill=tk.X, padx=10, pady=(5,10))
    
    # ==================== DIALOGS ====================
    
    def show_about(self):
        """Mostra finestra About"""
        about = tk.Toplevel(self)
        about.title("About")
        about.resizable(False, False)
        about.geometry("400x300")
        
        about.update_idletasks()
        x = (about.winfo_screenwidth() // 2) - (about.winfo_width() // 2)
        y = (about.winfo_screenheight() // 2) - (about.winfo_height() // 2)
        about.geometry(f"+{x}+{y}")
        
        frame = ttk.Frame(about, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=APP_TITLE, font=("Segoe UI", 14, "bold")).pack(pady=(10,5))
        ttk.Label(frame, text=f"Versione {APP_VERSION}", font=("Segoe UI", 10)).pack(pady=5)
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        ttk.Label(frame, text="Sviluppato da:", font=("Segoe UI", 9, "bold")).pack(pady=(5,2))
        ttk.Label(frame, text=APP_AUTHOR, font=("Segoe UI", 10)).pack(pady=2)
        ttk.Label(frame, text=APP_DEPT, font=("Segoe UI", 9), foreground="#666").pack(pady=2)
        ttk.Label(frame, text=APP_EMAIL, font=("Segoe UI", 9), foreground=COLORS['link']).pack(pady=2)
        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        info_text = ("Tool per l'analisi di tank FST/BBT con calcolo\n"
                    "estratto, variazioni di livello e grafici temporali.")
        ttk.Label(frame, text=info_text, font=("Segoe UI", 8), foreground=COLORS['info'], justify=tk.CENTER).pack(pady=5)
        
        ttk.Button(frame, text="Chiudi", command=about.destroy).pack(pady=(15,10))
    
    def show_formulas(self):
        """Mostra finestra formule (implementazione completa nel codice originale)"""
        messagebox.showinfo("Formule", 
            f"Formula principale:\n{FA_FORMULA}\n\n"
            "Kg estratto = f(A) × Level\n\n"
            "Vedi documentazione completa per dettagli.")
    
    # ==================== EVENTI PRINCIPALI ====================
    
    def on_open(self):
        """Apri file CSV"""
        path = filedialog.askopenfilename(
            title="Seleziona CSV",
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            self.analyzer = TankAnalyzer(path)
        except Exception as e:
            messagebox.showerror("Errore", str(e))
            return
        
        self.current_file = path
        self.lbl_file.config(text=os.path.basename(path))
        self.populate_days()
        self.populate_raw_data()
        self.on_apply()
    
    def populate_days(self):
        """Popola lista giorni"""
        self.days_list = []
        if not self.analyzer or not self.analyzer.rows or self.analyzer.time_idx is None:
            self.cb_day['values'] = []
            return
        
        seen = set()
        for r in self.analyzer.rows:
            if self.analyzer.time_idx < len(r):
                dt = parse_time(r[self.analyzer.time_idx])
                if dt:
                    dstr = dt.date().strftime("%Y-%m-%d")
                    if dstr not in seen:
                        seen.add(dstr)
                        self.days_list.append(dstr)
        
        self.days_list.sort()
        self.cb_day['values'] = self.days_list
        if self.days_list:
            self.sel_day.set(self.days_list[0])
    
    def on_select_day(self):
        """Giorno selezionato"""
        if self.sel_day.get().strip():
            self.on_apply()
    
    def on_prev_day(self):
        """Giorno precedente"""
        i = self._current_day_index()
        if i <= 0:
            return
        self.sel_day.set(self.days_list[i-1])
        self.on_select_day()
    
    def on_next_day(self):
        """Giorno successivo"""
        i = self._current_day_index()
        if i == -1 or i >= len(self.days_list)-1:
            return
        self.sel_day.set(self.days_list[i+1])
        self.on_select_day()
    
    def _current_day_index(self):
        """Indice giorno corrente"""
        if not self.days_list:
            return -1
        day = self.sel_day.get().strip()
        try:
            return self.days_list.index(day)
        except ValueError:
            return -1
    
    def on_apply(self):
        """Applica analisi"""
        if not self.analyzer:
            return
        
        day_str = self.sel_day.get().strip()
        if not day_str:
            messagebox.showwarning("Attenzione", "Seleziona un giorno dalla lista.")
            return
        
        t_from = self._parse_date(day_str)
        if t_from is None:
            messagebox.showerror("Errore", "Formato data non valido.")
            return
        
        t_to = t_from.replace(hour=23, minute=59, second=59)
        
        tanks, mats, debug = self.analyzer.analyze(
            t_from=t_from, t_to=t_to,
            include_fst=self.b_fst.get(), 
            include_bbt=self.b_bbt.get(),
            include_rbt=self.b_rbt.get()
        )
        
        self._cache_tank = tanks
        self._cache_mat = mats
        self._cache_debug = debug
        
        self._populate_summary_tables()
        self._populate_debug_table()
        self.update_variations_table()
        self.refresh_total_window()
    
    def _populate_summary_tables(self):
        """Popola tabelle riepilogo"""
        # Materiali
        for r in self.tv_mat.get_children():
            self.tv_mat.delete(r)
        for m, kg, fa, n in self._cache_mat:
            self.tv_mat.insert("", tk.END, values=(m, fmt_it(kg, 3), fmt_it(fa), n))
        
        # Tank
        for r in self.tv.get_children():
            self.tv.delete(r)
        for tank, mat, g_last, v_last, sum_fa, kg_ext, n in self._cache_tank:
            self.tv.insert("", tk.END, values=(
                tank,
                mat if mat else '',
                fmt_it(g_last, 2) if g_last is not None else "",
                fmt_it(v_last, 2) if v_last is not None else "",
                fmt_it(sum_fa),
                fmt_it(kg_ext, 3),
                n
            ))
    
    def _populate_debug_table(self):
        """Popola tabella debug"""
        for r in self.tv_debug.get_children():
            self.tv_debug.delete(r)
        
        debug_total = 0.0
        for dt, tank, mat, g, v, fa, kg in self._cache_debug:
            self.tv_debug.insert("", tk.END, values=(
                dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                tank, mat,
                fmt_it(g, 2) if g is not None else "",
                fmt_it(v, 2) if v is not None else "",
                fmt_it(fa, 6) if fa is not None else "",
                fmt_it(kg, 3)
            ))
            debug_total += kg
        
        self.lbl_debug_total.config(text=f"Totale Kg estratto (debug): {fmt_it(debug_total, 3)} | Righe: {len(self._cache_debug)}")
    
    def _parse_date(self, s):
        """Parse data YYYY-MM-DD"""
        s = str(s).strip()
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None
    
    # ==================== VARIAZIONI ====================
    
    def load_all_variations(self):
        """Carica le variazioni giornaliere per tutti i giorni del CSV"""
        if not self.analyzer:
            messagebox.showwarning("Attenzione", "Carica prima un file CSV")
            return
        
        # Mostra stato filtri
        print(f"[DEBUG] Filtri attivi: FST={self.b_fst.get()}, BBT={self.b_bbt.get()}, RBT={self.b_rbt.get()}")
        
        # Analizza tutti i giorni disponibili
        print("[DEBUG] Caricamento variazioni giornaliere...")
        
        # Import necessari
        from utils import to_float, calculate_fA, sanitize_level, normalize_material, is_valid_value
        
        # Raggruppa tutti i dati per giorno e tank
        daily_by_tank = defaultdict(lambda: defaultdict(list))
        
        for row in self.analyzer.rows:
            # Ottieni timestamp
            if self.analyzer.time_idx is None or self.analyzer.time_idx >= len(row):
                continue
            dt = parse_time(row[self.analyzer.time_idx])
            if dt is None:
                continue
            
            day_key = dt.date().strftime("%Y-%m-%d")
            
            for idx, tank_key, family in self.analyzer.avg_cols:
                # Filtro famiglia
                if family == 'FST' and not self.b_fst.get():
                    continue
                if family == 'BBT' and not self.b_bbt.get():
                    continue
                if family == 'RBT' and not self.b_rbt.get():
                    continue
                
                # DEBUG: mostra cosa viene processato
                # print(f"[DEBUG] Processando {tank_key} (famiglia: {family})")
                
                if idx >= len(row):
                    continue
                
                # Estrai gravity
                gravity = to_float(row[idx])
                if not is_valid_value(gravity):
                    continue
                
                fa_value = calculate_fA(gravity)
                if not is_valid_value(fa_value):
                    continue
                
                # Level
                level_idx = self.analyzer.level_idx.get(tank_key)
                level = to_float(row[level_idx]) if (level_idx is not None and level_idx < len(row)) else None
                level = sanitize_level(level)
                
                # Material
                mat_idx = self.analyzer.material_idx.get(tank_key)
                material = row[mat_idx] if (mat_idx is not None and mat_idx < len(row)) else None
                material = normalize_material(material)
                
                kg_extracted = fa_value * level
                
                daily_by_tank[tank_key][day_key].append((dt, material, gravity, level, fa_value, kg_extracted))
        
        print(f"[DEBUG] Tank trovati: {len(daily_by_tank)}")
        
        # Per ogni tank, prendi l'ultima misurazione di ogni giorno
        tank_daily_last = {}
        all_days = set()
        
        for tank, days_data in daily_by_tank.items():
            tank_daily_last[tank] = {}
            for day, measurements in days_data.items():
                all_days.add(day)
                # Ordina per timestamp e prendi l'ultimo
                measurements.sort(key=lambda x: x[0])
                last = measurements[-1]
                _, material, gravity, level, fa_value, kg_extracted = last
                tank_daily_last[tank][day] = (material, gravity, level, fa_value, kg_extracted)
        
        # Ordina i giorni
        sorted_days = sorted(all_days)
        print(f"[DEBUG] Giorni trovati: {len(sorted_days)}")
        
        # Calcola variazioni giorno per giorno
        self._cache_variations = []
        
        for tank in sorted(tank_daily_last.keys()):
            for i in range(1, len(sorted_days)):
                prev_day = sorted_days[i-1]
                curr_day = sorted_days[i]
                
                if prev_day not in tank_daily_last[tank] or curr_day not in tank_daily_last[tank]:
                    continue
                
                mat_prev, g_prev, v_prev, fa_prev, kg_prev = tank_daily_last[tank][prev_day]
                mat_curr, g_curr, v_curr, fa_curr, kg_curr = tank_daily_last[tank][curr_day]
                
                delta_level = v_curr - v_prev if (v_curr is not None and v_prev is not None) else None
                delta_kg = kg_curr - kg_prev
                
                self._cache_variations.append((
                    curr_day, tank, mat_curr, 
                    v_prev, v_curr, delta_level,
                    g_prev, g_curr,
                    kg_prev, kg_curr, delta_kg
                ))
        
        print(f"[DEBUG] Variazioni calcolate: {len(self._cache_variations)}")
        
        if not self._cache_variations:
            messagebox.showinfo("Info", "Nessuna variazione trovata. Verifica che ci siano almeno 2 giorni consecutivi con dati.")
        else:
            messagebox.showinfo("Successo", f"Caricate {len(self._cache_variations)} variazioni giornaliere!\n\n"
                              f"Giorni analizzati: {len(sorted_days)}\n"
                              f"Confronti giorno-giorno: {len(sorted_days)-1}")
        
        self.update_variations_table()
    
    def update_variations_table(self):
        """Aggiorna tabella variazioni con i dati calcolati"""
        # Pulisci tabella
        for r in self.tv_variations.get_children():
            self.tv_variations.delete(r)
        
        if not hasattr(self, '_cache_variations') or not self._cache_variations:
            self.lbl_var_summary.config(text="Nessuna variazione caricata. Clicca 'Carica Variazioni'.")
            return
        
        # Aggiorna lista tank disponibili
        all_tanks = sorted(set(v[1] for v in self._cache_variations))
        self.cb_filter_tank['values'] = ["Tutti"] + all_tanks
        
        # Filtra per tank selezionato
        filter_tank = self.var_filter_tank.get()
        if filter_tank != "Tutti":
            variations_to_show = [v for v in self._cache_variations if v[1] == filter_tank]
        else:
            variations_to_show = self._cache_variations
        
        # Statistiche
        total_delta_level = 0.0
        total_delta_kg = 0.0
        max_increase_level = 0.0
        max_decrease_level = 0.0
        max_increase_kg = 0.0
        max_decrease_kg = 0.0
        
        for var in variations_to_show:
            curr_day, tank, mat_curr, v_prev, v_curr, delta_level, g_prev, g_curr, kg_prev, kg_curr, delta_kg = var
            
            if delta_level is not None:
                total_delta_level += delta_level
                max_increase_level = max(max_increase_level, delta_level)
                max_decrease_level = min(max_decrease_level, delta_level)
            
            total_delta_kg += delta_kg
            max_increase_kg = max(max_increase_kg, delta_kg)
            max_decrease_kg = min(max_decrease_kg, delta_kg)
            
            # Tag per colore
            tag = ""
            if delta_level is not None:
                if delta_level < -THRESHOLDS['significant_level_change']:
                    tag = "decrease"
                elif delta_level > THRESHOLDS['significant_level_change']:
                    tag = "increase"
            
            self.tv_variations.insert("", tk.END, values=(
                curr_day, tank, mat_curr,
                fmt_it(v_prev, 2) if v_prev is not None else "",
                fmt_it(v_curr, 2) if v_curr is not None else "",
                fmt_it(delta_level, 2) if delta_level is not None else "",
                fmt_it(g_prev, 2) if g_prev is not None else "",
                fmt_it(g_curr, 2) if g_curr is not None else "",
                fmt_it(kg_prev, 3), fmt_it(kg_curr, 3), fmt_it(delta_kg, 3)
            ), tags=(tag,))
        
        num_variations = len(variations_to_show)
        
        if num_variations > 0:
            summary = (
                f"Variazioni giornaliere: {num_variations} | "
                f"ΔLevel totale: {fmt_it(total_delta_level, 2)} hl | "
                f"ΔKg totale: {fmt_it(total_delta_kg, 3)} kg\n"
                f"Max aumento level: {fmt_it(max_increase_level, 2)} hl | "
                f"Max calo level: {fmt_it(max_decrease_level, 2)} hl | "
                f"Max aumento Kg: {fmt_it(max_increase_kg, 3)} kg | "
                f"Max perdita Kg: {fmt_it(max_decrease_kg, 3)} kg"
            )
        else:
            summary = "Nessuna variazione disponibile"
        
        self.lbl_var_summary.config(text=summary)
        print(f"[DEBUG] Variazioni visualizzate: {num_variations}")
    
    # ==================== DATI RAW ====================
    
    def populate_raw_data(self):
        """Popola tabella dati raw"""
        if not self.analyzer:
            return
        
        for item in self.tv_raw.get_children():
            self.tv_raw.delete(item)
        
        self.tv_raw['columns'] = self.analyzer.header
        
        for col in self.analyzer.header:
            self.tv_raw.heading(col, text=col)
            col_width = min(max(len(col) * 8, 80), 200)
            self.tv_raw.column(col, width=col_width, anchor=tk.W)
        
        for row in self.analyzer.rows:
            padded_row = row + [''] * (len(self.analyzer.header) - len(row))
            self.tv_raw.insert("", tk.END, values=padded_row[:len(self.analyzer.header)])
        
        self.lbl_raw_info.config(
            text=f"Righe totali: {len(self.analyzer.rows)} | Colonne: {len(self.analyzer.header)} | "
                 f"Periodo: {self.analyzer.min_time.strftime('%Y-%m-%d') if self.analyzer.min_time else 'N/A'} - "
                 f"{self.analyzer.max_time.strftime('%Y-%m-%d') if self.analyzer.max_time else 'N/A'}"
        )
    
    # ==================== GRAFICI ====================
    
    def on_generate_charts(self):
        """Genera grafici"""
        if not self.analyzer or not HAS_MATPLOTLIB:
            return
        
        for widget in self.charts_frame.winfo_children():
            widget.destroy()
        
        self.chart_figures = []
        
        daily_data = self.analyzer.analyze_all_days(
            include_fst=self.b_fst.get(),
            include_bbt=self.b_bbt.get(),
            include_rbt=self.b_rbt.get()
        )
        
        if not daily_data:
            ttk.Label(self.charts_frame, text="Nessun dato disponibile per i grafici").pack(pady=20)
            return
        
        days = sorted(daily_data.keys())
        kg_totals = [daily_data[d]['kg'] for d in days]
        
        # Grafico 1: Totale
        fig1 = Figure(figsize=(12, 4), dpi=80)
        ax1 = fig1.add_subplot(111)
        ax1.plot(days, kg_totals, marker='o', linewidth=2, markersize=6)
        ax1.set_title('Kg Estratto Totale per Giorno')
        ax1.set_xlabel('Giorno')
        ax1.set_ylabel('Kg Estratto')
        ax1.grid(True, alpha=0.3)
        ax1.tick_params(axis='x', rotation=45)
        fig1.tight_layout()
        self.chart_figures.append(fig1)
        
        canvas1 = FigureCanvasTkAgg(fig1, self.charts_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Grafico 2: Per materiale (top 5)
        all_materials = set()
        for d in days:
            all_materials.update(daily_data[d]['by_material'].keys())
        
        material_totals = {}
        for mat in all_materials:
            material_totals[mat] = sum(daily_data[d]['by_material'].get(mat, 0) for d in days)
        
        top_materials = sorted(material_totals.items(), key=lambda x: -x[1])[:5]
        
        if top_materials:
            fig2 = Figure(figsize=(12, 4), dpi=80)
            ax2 = fig2.add_subplot(111)
            
            for mat, _ in top_materials:
                values = [daily_data[d]['by_material'].get(mat, 0) for d in days]
                ax2.plot(days, values, marker='o', label=mat, linewidth=2, markersize=5)
            
            ax2.set_title('Kg Estratto per Materiale (Top 5)')
            ax2.set_xlabel('Giorno')
            ax2.set_ylabel('Kg Estratto')
            ax2.legend(loc='best')
            ax2.grid(True, alpha=0.3)
            ax2.tick_params(axis='x', rotation=45)
            fig2.tight_layout()
            self.chart_figures.append(fig2)
            
            canvas2 = FigureCanvasTkAgg(fig2, self.charts_frame)
            canvas2.draw()
            canvas2.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Grafico 3: Per tank (top 5)
        all_tanks = set()
        for d in days:
            all_tanks.update(daily_data[d]['by_tank'].keys())
        
        tank_totals = {}
        for tank in all_tanks:
            tank_totals[tank] = sum(daily_data[d]['by_tank'].get(tank, 0) for d in days)
        
        top_tanks = sorted(tank_totals.items(), key=lambda x: -x[1])[:5]
        
        if top_tanks:
            fig3 = Figure(figsize=(12, 4), dpi=80)
            ax3 = fig3.add_subplot(111)
            
            for tank, _ in top_tanks:
                values = [daily_data[d]['by_tank'].get(tank, 0) for d in days]
                ax3.plot(days, values, marker='o', label=tank, linewidth=2, markersize=5)
            
            ax3.set_title('Kg Estratto per Tank (Top 5)')
            ax3.set_xlabel('Giorno')
            ax3.set_ylabel('Kg Estratto')
            ax3.legend(loc='best')
            ax3.grid(True, alpha=0.3)
            ax3.tick_params(axis='x', rotation=45)
            fig3.tight_layout()
            self.chart_figures.append(fig3)
            
            canvas3 = FigureCanvasTkAgg(fig3, self.charts_frame)
            canvas3.draw()
            canvas3.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=5)
    
    def on_export_charts_pdf(self):
        """Esporta grafici in PDF"""
        if not self.chart_figures:
            messagebox.showwarning("Attenzione", "Genera prima i grafici cliccando 'Genera Grafici'")
            return
        
        if not HAS_MATPLOTLIB:
            messagebox.showerror("Errore", "Matplotlib non disponibile")
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva Grafici PDF",
            defaultextension=".pdf",
            initialfile=EXPORT_FILENAMES['pdf'],
            filetypes=[("PDF", "*.pdf"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            with PdfPages(path) as pdf:
                for fig in self.chart_figures:
                    pdf.savefig(fig, bbox_inches='tight')
                
                # Pagina info
                fig_info = Figure(figsize=(8.5, 11))
                ax_info = fig_info.add_subplot(111)
                ax_info.axis('off')
                
                info_text = (
                    f"{APP_TITLE}\n\n"
                    f"Report Grafici Analisi Tank\n\n"
                    f"File: {os.path.basename(self.current_file) if self.current_file else 'N/A'}\n"
                    f"Data generazione: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"Grafici generati: {len(self.chart_figures)}\n\n"
                    f"FST inclusi: {'Sì' if self.b_fst.get() else 'No'}\n"
                    f"BBT inclusi: {'Sì' if self.b_bbt.get() else 'No'}\n\n"
                    f"──────────────────────────────\n\n"
                    f"Sviluppato da: {APP_AUTHOR}\n"
                    f"Dipartimento: {APP_DEPT}\n"
                    f"Versione: {APP_VERSION}\n"
                    f"Contatto: {APP_EMAIL}"
                )
                
                ax_info.text(0.5, 0.5, info_text, ha='center', va='center',
                           fontsize=12, family='monospace', transform=ax_info.transAxes)
                
                pdf.savefig(fig_info, bbox_inches='tight')
            
            messagebox.showinfo("Esportato", f"Grafici salvati in PDF:\n{path}\n\nPagine: {len(self.chart_figures) + 1}")
        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante l'esportazione PDF:\n{str(e)}")
    
    # ==================== TOTALE CANTINA ====================
    
    def compute_totals(self):
        """Calcola totali cantina"""
        if not self._cache_mat:
            return 0.0, 0.0, 0
        
        excl0 = self.var_exclude_mat0.get()
        tot_kg = tot_fa = tot_n = 0
        
        for m, kg, fa, n in self._cache_mat:
            # m è già normalizzato, quindi controlla sia '0' che 'vuoto'
            if excl0 and (str(m).strip() == '0' or str(m).strip().lower() == 'vuoto'):
                continue
            tot_kg += kg
            tot_fa += fa
            tot_n += n
        
        return tot_kg, tot_fa, tot_n
    
    def on_show_total(self):
        """Mostra finestra totale cantina"""
        if self._tot_win and tk.Toplevel.winfo_exists(self._tot_win):
            self._tot_win.lift()
            self.refresh_total_window()
            return
        
        self._tot_win = tk.Toplevel(self)
        self._tot_win.title("Totale Cantina")
        self._tot_win.resizable(False, False)
        
        frm = ttk.Frame(self._tot_win, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)
        
        self.lbl_tot_kg = ttk.Label(frm, text="Totale Kg estratto: -", font=("Segoe UI", 10, "bold"))
        self.lbl_tot_fa = ttk.Label(frm, text="Totale f(A): -")
        self.lbl_tot_n = ttk.Label(frm, text="Misure conteggiate: -")
        self.lbl_note = ttk.Label(frm, text="", foreground=COLORS['info'])
        
        self.lbl_tot_kg.pack(anchor=tk.W)
        self.lbl_tot_fa.pack(anchor=tk.W, pady=(4,0))
        self.lbl_tot_n.pack(anchor=tk.W, pady=(4,6))
        self.lbl_note.pack(anchor=tk.W)
        
        btns = ttk.Frame(frm)
        btns.pack(fill=tk.X, pady=(10,0))
        ttk.Button(btns, text="Chiudi", command=self._tot_win.destroy).pack(side=tk.RIGHT)
        
        self.refresh_total_window()
    
    def refresh_total_window(self):
        """Aggiorna finestra totale"""
        if not (self._tot_win and tk.Toplevel.winfo_exists(self._tot_win)):
            return
        
        kg, fa, n = self.compute_totals()
        self.lbl_tot_kg.config(text=f"Totale Kg estratto: {fmt_it(kg, 3)}")
        self.lbl_tot_fa.config(text=f"Totale f(A): {fmt_it(fa)}")
        self.lbl_tot_n.config(text=f"Misure conteggiate: {n}")
        
        note = "(Material=0 escluso)" if self.var_exclude_mat0.get() else "(Material=0 incluso)"
        self.lbl_note.config(text=note)
    
    # ==================== EXPORT ====================
    
    def on_export_tank_csv(self):
        """Esporta per tank CSV"""
        if not self._cache_tank:
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva CSV (per Tank)",
            defaultextension=".csv",
            initialfile=EXPORT_FILENAMES['tank_csv'],
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Tank','Materiale','Gravity_ultimo','Volume_ultimo','Somma_f(A)','Kg_estratto','Misure'])
                for tank, mat, g_last, v_last, sum_fa, kg_ext, n in self._cache_tank:
                    w.writerow([
                        tank, mat or '',
                        f"{g_last:.2f}" if g_last is not None else "",
                        f"{v_last:.2f}" if v_last is not None else "",
                        f"{sum_fa:.6f}", f"{kg_ext:.3f}", n
                    ])
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
    
    def on_export_mat_csv(self):
        """Esporta per materiale CSV"""
        if not self._cache_mat:
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva CSV (per Materiale)",
            defaultextension=".csv",
            initialfile=EXPORT_FILENAMES['material_csv'],
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Materiale','Kg_estratto','Somma_f(A)','Misure'])
                for m, kg, fa, n in self._cache_mat:
                    w.writerow([m, f"{kg:.3f}", f"{fa:.6f}", n])
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
    
    def on_export_debug_csv(self):
        """Esporta debug CSV"""
        if not self._cache_debug:
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva CSV Debug",
            defaultextension=".csv",
            initialfile=EXPORT_FILENAMES['debug_csv'],
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Timestamp','Tank','Materiale','Gravity','Level_hl','f(A)','Kg_estratto'])
                for dt, tank, mat, g, v, fa, kg in self._cache_debug:
                    w.writerow([
                        dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                        tank, mat,
                        f"{g:.2f}" if g is not None else "",
                        f"{v:.2f}" if v is not None else "",
                        f"{fa:.6f}" if fa is not None else "",
                        f"{kg:.3f}"
                    ])
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
    
    def on_export_variations_csv(self):
        """Esporta variazioni CSV"""
        if not hasattr(self, '_cache_variations') or not self._cache_variations:
            messagebox.showwarning("Attenzione", "Carica prima le variazioni cliccando 'Carica Variazioni'")
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva CSV Variazioni",
            defaultextension=".csv",
            initialfile=EXPORT_FILENAMES['variations_csv'],
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(['Data','Tank','Materiale','Level_Prec_hl','Level_Corr_hl','Delta_Level_hl',
                           'Gravity_Prec','Gravity_Corr','Kg_Prec','Kg_Corr','Delta_Kg'])
                
                for var in self._cache_variations:
                    curr_day, tank, mat_curr, v_prev, v_curr, delta_level, g_prev, g_curr, kg_prev, kg_curr, delta_kg = var
                    
                    w.writerow([
                        curr_day, tank, mat_curr,
                        f"{v_prev:.2f}" if v_prev is not None else "",
                        f"{v_curr:.2f}" if v_curr is not None else "",
                        f"{delta_level:.2f}" if delta_level is not None else "",
                        f"{g_prev:.2f}" if g_prev is not None else "",
                        f"{g_curr:.2f}" if g_curr is not None else "",
                        f"{kg_prev:.3f}", f"{kg_curr:.3f}", f"{delta_kg:.3f}"
                    ])
            
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}\n\nVariazioni: {len(self._cache_variations)}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
    
    def on_export_raw_csv(self):
        """Esporta dati raw CSV"""
        if not self.analyzer:
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva CSV Raw",
            defaultextension=".csv",
            initialfile=EXPORT_FILENAMES['raw_csv'],
            filetypes=[("CSV", "*.csv"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                w.writerow(self.analyzer.header)
                w.writerows(self.analyzer.rows)
            
            messagebox.showinfo("Esportato", f"File raw salvato in:\n{path}\n\nRighe: {len(self.analyzer.rows)}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
    
    def on_export_xlsx(self):
        """Esporta report XLSX completo"""
        if not self._cache_mat and not self._cache_tank:
            return
        if not HAS_OPENPYXL:
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva XLSX",
            defaultextension=".xlsx",
            initialfile=EXPORT_FILENAMES['xlsx'],
            filetypes=[("Excel", "*.xlsx"), ("Tutti i file", "*.*")]
        )
        if not path:
            return
        
        try:
            wb = Workbook()
            
            # Foglio 1: Per Materiale
            ws1 = wb.active
            ws1.title = 'Per Materiale'
            ws1.append(['Materiale','Kg_estratto','Somma_f(A)','Misure'])
            for m, kg, fa, n in self._cache_mat:
                ws1.append([m, float(f"{kg:.3f}"), float(f"{fa:.6f}"), n])
            
            # Foglio 2: Per Tank
            ws2 = wb.create_sheet('Per Tank')
            ws2.append(['Tank','Materiale','Gravity_ultimo','Volume_ultimo','Somma_f(A)','Kg_estratto','Misure'])
            for tank, mat, g_last, v_last, sum_fa, kg_ext, n in self._cache_tank:
                ws2.append([
                    tank, mat or '',
                    float(f"{g_last:.2f}") if g_last is not None else None,
                    float(f"{v_last:.2f}") if v_last is not None else None,
                    float(f"{sum_fa:.6f}"), float(f"{kg_ext:.3f}"), n
                ])
            
            # Foglio 3: Debug
            ws3 = wb.create_sheet('Debug')
            ws3.append(['Timestamp','Tank','Materiale','Gravity','Level_hl','f(A)','Kg_estratto'])
            for dt, tank, mat, g, v, fa, kg in self._cache_debug:
                ws3.append([
                    dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
                    tank, mat,
                    float(f"{g:.2f}") if g is not None else None,
                    float(f"{v:.2f}") if v is not None else None,
                    float(f"{fa:.6f}") if fa is not None else None,
                    float(f"{kg:.3f}")
                ])
            
            # Foglio 4: Note
            ws4 = wb.create_sheet('Note')
            ws4.append(['Descrizione','Valore'])
            ws4.append(['Equivalenza', "'Average Gravity' == 'Average Plato' (usati come 'Gravity')"])
            ws4.append(['f(A)', FA_FORMULA])
            ws4.append(['Kg estratto (riga)', 'f(A) * Level'])
            ws4.append(['Aggregazioni', 'Somme su periodo filtrato; Material per riga secondo colonna Material del tank'])
            ws4.append(['Mapping Material', '7=ichnusa; 8=non filtrata; 9=cruda; 28=ambra limpida'])
            excl = 'sì' if self.var_exclude_mat0.get() else 'no'
            ws4.append(['Totale Cantina', f"Material=0 escluso: {excl}"])
            ws4.append(['Modalità', 'Giorno singolo'])
            ws4.append(['', ''])
            ws4.append(['Tool Info', ''])
            ws4.append(['Sviluppato da', APP_AUTHOR])
            ws4.append(['Dipartimento', APP_DEPT])
            ws4.append(['Versione', APP_VERSION])
            ws4.append(['Contatto', APP_EMAIL])
            
            wb.save(path)
            messagebox.showinfo("Esportato", f"File salvato in:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))


# ==================== MAIN ====================

if __name__ == '__main__':
    app = TankAnalysisApp()
    app.mainloop()